import openpyxl
from openpyxl import Workbook

wb = openpyxl.load_workbook("Book1.xlsx")   # address of Book 1 is in python directory
sheets = wb.sheetnames
sh1 = wb['Sheet1']

wb1 = Workbook()
wb1['Sheet'].title = "Report"
sh2 = wb1.active
sh2['A1'].value = "x"
sh2['B1'].value = "y"

for n in range(2, 46):
    x_old = sh1.cell(row=n, column=1).value
    y_old = sh1.cell(row=n, column=2).value
    sh2.cell(row=abs(n-47), column=1).value = x_old
    sh2.cell(row=abs(n-47), column=2).value = y_old

wb1.save("C:\\Users\\Book.xlsx")    # address of destination file
